from openpyxl import Workbook
from openpyxl.styles import Alignment

# Create a new workbook and select the active sheet
wb = Workbook()
sheet = wb.active
sheet.title = "Merged Cells"

# Add data to the sheet
sheet["A1"] = "This is a merged cell"
sheet.merge_cells("A1:D1")  # Merge cells from A1 to D1

# Center align the text in the merged cell (optional)
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Merge cells across rows
sheet["A2"] = "Merged across rows"
sheet.merge_cells("A2:A4")  # Merge cells from A2 to A4
sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

# Save the workbook
wb.save("merged_cells.xlsx")
