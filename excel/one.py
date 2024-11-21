from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook

# Create a new workbook
wb0 = Workbook()
sheet0 = wb0.active
sheet0.title = "Instructions"

wb = Workbook()
sheet = wb.active
sheet.title = "Patient Log"

# Apply styles
header_font = Font(bold=True, color="FFF22F")
header_fill = PatternFill(start_color="bfbfbf", end_color="bfbfbf", fill_type="solid")
center_align = Alignment(horizontal="center")

required_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
optional_fill = PatternFill(start_color="a6a6a6", end_color="a6a6a6", fill_type="solid")

black_text = Font(color="000000")
white_text = Font(color="ffffff")


# Style the header
sheet["A1"].font = header_font
sheet["A1"].fill = header_fill
sheet["A1"].alignment = center_align
sheet["B1"].font = header_font
sheet["B1"].fill = header_fill
sheet["B1"].alignment = center_align
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["A2"] = "John"
sheet["B2"] = 30
sheet["A3"] = "Destin"
sheet["B3"] = 25
sheet["A4"] = "Delphin"
sheet["B4"] = 23

wb.save("styled_example.xlsx")
