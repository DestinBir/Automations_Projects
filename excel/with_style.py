from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Page 1"

# Apply styles
header_font = Font(bold=True, color="FFF22F")
header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
center_align = Alignment(horizontal="center")

# Style the header
sheet["A1"].font = header_font
sheet["A1"].fill = header_fill
sheet["A1"].alignment = center_align
sheet["B1"].font = header_font
sheet["B1"].fill = header_fill
sheet["B1"].alignment = center_align
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["A2"] = "John"
sheet["B2"] = 30
sheet["A3"] = "Destin"
sheet["B3"] = 25
sheet["A4"] = "Delphin"
sheet["B4"] = 23

wb.save("styled_example.xlsx")
