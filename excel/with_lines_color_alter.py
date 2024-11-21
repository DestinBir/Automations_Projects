from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a workbook and select the active worksheet
wb = Workbook()
sheet = wb.active
sheet.title = "Alternate Line Colors"

# Sample data
data = [
    ["Name", "Age", "City"],
    ["Alice", 25, "New York"],
    ["Bob", 30, "London"],
    ["Charlie", 35, "Paris"],
    ["Diana", 28, "Berlin"],
    ["Edward", 40, "Tokyo"],
]

# Write data to the sheet
for row_idx, row in enumerate(data, start=1):
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# Define alternating row colors
light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
dark_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")  # Dark green

# Apply alternate colors (skipping the header row)
for row_idx in range(2, sheet.max_row + 1):  # Start at the second row
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        if row_idx % 2 == 0:  # Even row
            cell.fill = light_fill
        else:  # Odd row
            cell.fill = dark_fill

# Save the workbook
wb.save("alternate_colors.xlsx")
